"""
步骤1：歌曲分类核心功能
从原始的 song_classifier_gui.py 中提取
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Callable, Optional, Tuple, List
import re
from datetime import datetime, timedelta, time


# 调号映射：字母调号到数字的转换
KEY_MAPPING = {
    'C': 0, 'C#': 1, 'Db': 1,
    'D': 2, 'D#': 3, 'Eb': 3,
    'E': 4,
    'F': 5, 'F#': 6, 'Gb': 6,
    'G': 7, 'G#': 8, 'Ab': 8,
    'A': 9, 'A#': 10, 'Bb': 10,
    'B': 11
}

# 数字到字母调号的映射
NUM_TO_KEY = {
    0: 'C', 1: 'C#', 2: 'D', 3: 'D#', 4: 'E',
    5: 'F', 6: 'F#', 7: 'G', 8: 'G#', 9: 'A',
    10: 'A#', 11: 'B'
}


def key_to_number(key):
    """将调号转换为数字"""
    if pd.isna(key):
        raise ValueError("调号不能为空")
    
    key_str = str(key).strip()
    
    # 如果是数字，直接转换
    try:
        num = int(float(key_str))
        return num % 12
    except ValueError:
        pass
    
    # 尝试字母映射
    if key_str in KEY_MAPPING:
        return KEY_MAPPING[key_str]
    
    # 大小写不敏感匹配
    key_upper = key_str.upper()
    for k, v in KEY_MAPPING.items():
        if k.upper() == key_upper:
            return v
    
    raise ValueError(f"无法识别的调号: {key_str}")


def number_to_key(num):
    """将数字调号转换为字母调号"""
    num = int(num) % 12
    return NUM_TO_KEY.get(num, str(num))


def format_time_to_mmssmmm(time_value) -> str:
    """
    将各种时间格式统一转换为 MM:SS.mmm 格式 (例如: 01:17.877)
    
    Args:
        time_value: 时间值，可能是字符串、datetime对象、timedelta、浮点数等
    
    Returns:
        格式化后的时间字符串，格式为 MM:SS.mmm
    """
    if pd.isna(time_value) or time_value == '' or time_value is None:
        return ''
    
    # 如果已经是正确格式的字符串，直接返回
    if isinstance(time_value, str):
        time_str = str(time_value).strip()
        # 检查是否符合 MM:SS.mmm 格式
        if re.match(r'^\d{2}:\d{2}\.\d{3}$', time_str):
            return time_str
        # 检查是否符合 M:SS.mmm 格式（单数字分钟）
        if re.match(r'^\d{1,2}:\d{2}\.\d{1,3}$', time_str):
            parts = time_str.split(':')
            minutes = parts[0].zfill(2)
            seconds_part = parts[1]
            # 确保毫秒部分是3位
            if '.' in seconds_part:
                sec, ms = seconds_part.split('.')
                ms = ms[:3].ljust(3, '0')
                return f"{minutes}:{sec}.{ms}"
            else:
                return f"{minutes}:{seconds_part}.000"
    
    # 如果是datetime.time对象
    if isinstance(time_value, time):
        total_seconds = time_value.hour * 3600 + time_value.minute * 60 + time_value.second
        milliseconds = time_value.microsecond // 1000
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes:02d}:{seconds:02d}.{milliseconds:03d}"
    
    # 如果是datetime对象
    if isinstance(time_value, datetime):
        total_seconds = time_value.hour * 3600 + time_value.minute * 60 + time_value.second
        milliseconds = time_value.microsecond // 1000
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes:02d}:{seconds:02d}.{milliseconds:03d}"
    
    # 如果是timedelta对象
    if isinstance(time_value, timedelta):
        total_seconds = int(time_value.total_seconds())
        milliseconds = int((time_value.total_seconds() - total_seconds) * 1000)
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes:02d}:{seconds:02d}.{milliseconds:03d}"
    
    # 如果是浮点数（可能是Excel中的时间格式，如0.052表示天数，或直接的秒数）
    if isinstance(time_value, (int, float)):
        try:
            # Excel时间格式：1 = 1天，0.5 = 12小时
            # 判断：如果值小于1且大于0，可能是Excel时间格式（天数）
            # 如果值大于等于1且小于86400，可能是秒数
            # 如果值大于等于86400，可能是毫秒数
            if 0 < time_value < 1.0:  # Excel时间格式（天数）
                total_seconds = time_value * 86400
            elif 1.0 <= time_value < 86400:  # 直接是秒数
                total_seconds = float(time_value)
            elif time_value >= 86400:  # 可能是毫秒数
                total_seconds = float(time_value) / 1000
            else:  # 其他情况，假设是秒数
                total_seconds = abs(float(time_value))
            
            minutes = int(total_seconds) // 60
            seconds = int(total_seconds) % 60
            milliseconds = int((total_seconds - int(total_seconds)) * 1000)
            return f"{minutes:02d}:{seconds:02d}.{milliseconds:03d}"
        except:
            pass
    
    # 尝试解析字符串格式
    time_str = str(time_value).strip()
    
    # 尝试解析 MM:SS.mmm 或 M:SS.mmm
    match = re.match(r'^(\d{1,2}):(\d{2})\.?(\d{0,3})$', time_str)
    if match:
        minutes = int(match.group(1))
        seconds = int(match.group(2))
        ms_str = match.group(3) if match.group(3) else '0'
        milliseconds = int(ms_str.ljust(3, '0')[:3])
        return f"{minutes:02d}:{seconds:02d}.{milliseconds:03d}"
    
    # 尝试解析 SS.mmm 格式（只有秒）
    match = re.match(r'^(\d+)\.?(\d{0,3})$', time_str)
    if match:
        total_seconds = int(match.group(1))
        ms_str = match.group(2) if match.group(2) else '0'
        milliseconds = int(ms_str.ljust(3, '0')[:3])
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes:02d}:{seconds:02d}.{milliseconds:03d}"
    
    # 如果无法解析，返回空字符串
    return ''


def classify_songs_core(
    excel_path: Path,
    output_path: Path,
    progress_callback: Optional[Callable[[int, int, str], bool]] = None,
    key_range: int = 2,
    bpm_range: int = 5
) -> Tuple[List[List[int]], pd.DataFrame]:
    """
    歌曲分类核心功能
    
    Args:
        excel_path: 输入的Excel文件路径
        output_path: 输出的Excel文件路径
        progress_callback: 进度回调函数 (当前进度, 总数, 消息) -> 是否继续
        key_range: 调号匹配区间（±几个半音），默认2
        bpm_range: 速度匹配区间（±几个BPM），默认5
    
    Returns:
        (groups, df) - 分类组列表和原始数据DataFrame
    """
    # 读取Excel文件（正常读取，时间格式将在后续格式化函数中处理）
    df = pd.read_excel(excel_path)
    
    # 检查必需的列
    has_song_name = '歌名' in df.columns or 'name' in df.columns
    has_key = '调号' in df.columns or 'key' in df.columns
    has_bpm = '速度' in df.columns or 'bpm' in df.columns
    
    if not (has_song_name and has_key and has_bpm):
        raise ValueError("Excel文件必须包含 '歌名'、'调号'、'速度' 三列")
    
    # 统一列名
    if 'name' in df.columns and '歌名' not in df.columns:
        df.rename(columns={'name': '歌名'}, inplace=True)
    if 'key' in df.columns and '调号' not in df.columns:
        df.rename(columns={'key': '调号'}, inplace=True)
    if 'bpm' in df.columns and '速度' not in df.columns:
        df.rename(columns={'bpm': '速度'}, inplace=True)
    
    # 处理可选列
    optional_columns = {
        'ID': 'id',
        'Chord Ai': 'Chord Ai',
        '歌手': '歌手',
        '副歌开始时间': '副歌开始时间',
        '副歌结束时间': '副歌结束时间',
        '段落剪切时间': '段落剪切时间',
        '性别': '性别'
    }
    
    # 统一列名
    for display_name, internal_name in optional_columns.items():
        if internal_name == 'id':
            if 'ID' in df.columns:
                df.rename(columns={'ID': 'id'}, inplace=True)
            elif 'id' not in df.columns:
                df['id'] = df.index + 1
        elif internal_name not in df.columns:
            df[internal_name] = ''
    
    # 处理性别列的不同写法
    gender_column = None
    for col_name in ['性别', 'gender', 'Gender', 'SEX', 'sex']:
        if col_name in df.columns:
            gender_column = col_name
            break
    if gender_column and gender_column != '性别':
        df.rename(columns={gender_column: '性别'}, inplace=True)
    elif '性别' not in df.columns:
        df['性别'] = ''
    
    # 处理歌手列
    if '歌手名' in df.columns:
        df.rename(columns={'歌手名': '歌手'}, inplace=True)
    elif 'artist' in df.columns:
        df.rename(columns={'artist': '歌手'}, inplace=True)
    elif '歌手' not in df.columns:
        df['歌手'] = ''
    
    # 规范化性别信息
    df['性别'] = df['性别'].fillna('').astype(str).str.strip()
    df['gender_normalized'] = df['性别'].str.lower()
    
    # 保存原始调号值
    df['调号_original'] = df['调号'].copy()
    
    # 转换调号为数字
    df['key_num'] = df['调号'].apply(key_to_number)
    
    # 转换速度为数值
    df['速度'] = pd.to_numeric(df['速度'], errors='coerce')
    
    if df['速度'].isna().any():
        raise ValueError("Excel文件中存在无效的速度值")
    
    # 分类处理
    groups = []
    used_as_anchor = set()
    
    total_songs = len(df)
    
    for anchor_idx in range(total_songs):
        current_group = [anchor_idx]
        used_as_anchor.add(anchor_idx)
        
        # 获取锚点信息
        anchor_key_num = df.iloc[anchor_idx]['key_num']
        anchor_bpm = df.iloc[anchor_idx]['速度']
        anchor_gender = df.iloc[anchor_idx]['gender_normalized']
        
        # 查找匹配歌曲
        for candidate_idx in range(total_songs):
            if candidate_idx in used_as_anchor:
                continue
            
            candidate_key_num = df.iloc[candidate_idx]['key_num']
            candidate_bpm = df.iloc[candidate_idx]['速度']
            candidate_gender = df.iloc[candidate_idx]['gender_normalized']
            
            # 性别不同跳过
            if candidate_gender != anchor_gender:
                continue
            
            # 检查调号差异（使用用户设定的区间）
            key_diff = abs(candidate_key_num - anchor_key_num)
            key_diff_circular = min(key_diff, 12 - key_diff)
            if key_diff_circular > key_range:
                continue
            
            # 检查速度差异（使用用户设定的区间）
            bpm_diff = abs(candidate_bpm - anchor_bpm)
            if bpm_diff <= bpm_range:
                current_group.append(candidate_idx)
        
        groups.append(current_group)
    
    # 生成Excel文件 - 所有分类结果放在同一个sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "分类结果"
    
    # 统计有效分类组数量（用于进度显示）
    valid_groups = [g for g in groups if len(g) > 1]
    total_valid_groups = len(valid_groups)
    
    # 定义列
    original_columns = ['ID', 'Chord Ai', '歌名', '歌手', '副歌开始时间', 
                      '副歌结束时间', '段落剪切时间', '调号', '速度', '性别']
    headers = original_columns + ['成品名']
    
    # 设置样式
    yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    num_original_cols = len(original_columns)
    for col_idx in range(1, num_original_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = headers[col_idx - 1]
        cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    product_col_idx = num_original_cols + 1
    cell = ws.cell(row=1, column=product_col_idx)
    cell.value = headers[-1]
    cell.fill = green_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    
    # 列映射
    column_mapping = {
        'ID': 'id',
        'Chord Ai': 'Chord Ai',
        '歌名': '歌名',
        '歌手': '歌手',
        '副歌开始时间': '副歌开始时间',
        '副歌结束时间': '副歌结束时间',
        '段落剪切时间': '段落剪切时间',
        '调号': '调号_original',
        '速度': '速度',
        '性别': '性别'
    }
    
    # 时间列名列表，用于格式化处理
    time_col_names = ['副歌开始时间', '副歌结束时间', '段落剪切时间']
    
    current_row = 2
    current_group_num = 0
    
    # 辅助函数：获取ID值
    def get_id_value(song_data):
        """从歌曲数据中获取ID值"""
        value = song_data.get('id', '')
        if value and not pd.isna(value):
            try:
                return int(value) if float(value) == int(float(value)) else str(value)
            except:
                return str(value)
        return ''
    
    for group_num, group_indices in enumerate(groups, start=1):
        if len(group_indices) == 1:
            continue
        
        current_group_num += 1
        
        # 进度回调：显示当前生成的分类数量
        if progress_callback:
            if not progress_callback(current_group_num, total_valid_groups, 
                                   f"生成分类 {current_group_num}/{total_valid_groups}"):
                break
        
        group_songs = df.iloc[group_indices].copy()
        anchor_song = group_songs.iloc[0]
        anchor_id = get_id_value(anchor_song)
        
        # 处理每对匹配歌曲
        for j in range(1, len(group_indices)):
            match_song = group_songs.iloc[j]
            match_id = get_id_value(match_song)
            
            # 生成成品名：ID1-ID2-拼接成品
            combined_name = f"{anchor_id}-{match_id}-拼接成品"
            
            # 第一行：锚定歌曲
            for col_idx, col_name in enumerate(original_columns, start=1):
                df_col_name = column_mapping.get(col_name, col_name)
                value = anchor_song.get(df_col_name, '')
                # 确保ID不是NaN，并转换为字符串或整数
                if col_name == 'ID' and value and not pd.isna(value):
                    try:
                        value = int(value) if float(value) == int(float(value)) else str(value)
                    except:
                        value = str(value)
                # 时间列格式化为 MM:SS.mmm 格式
                if col_name in time_col_names:
                    value = format_time_to_mmssmmm(value)
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                # 时间列设置为文本格式
                if col_name in time_col_names:
                    cell.number_format = '@'
            
            # 第二行：匹配歌曲
            for col_idx, col_name in enumerate(original_columns, start=1):
                df_col_name = column_mapping.get(col_name, col_name)
                value = match_song.get(df_col_name, '')
                # 确保ID不是NaN，并转换为字符串或整数
                if col_name == 'ID' and value and not pd.isna(value):
                    try:
                        value = int(value) if float(value) == int(float(value)) else str(value)
                    except:
                        value = str(value)
                # 时间列格式化为 MM:SS.mmm 格式
                if col_name in time_col_names:
                    value = format_time_to_mmssmmm(value)
                cell = ws.cell(row=current_row + 1, column=col_idx)
                cell.value = value
                # 时间列设置为文本格式
                if col_name in time_col_names:
                    cell.number_format = '@'
            
            # 合并成品名列
            product_col_letter = get_column_letter(product_col_idx)
            ws.merge_cells(f'{product_col_letter}{current_row}:{product_col_letter}{current_row + 1}')
            ws.cell(row=current_row, column=product_col_idx).value = combined_name
            merged_cell = ws.cell(row=current_row, column=product_col_idx)
            merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            merged_cell.border = thin_border
            
            # 设置边框
            for row in [current_row, current_row + 1]:
                for col in range(1, product_col_idx + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border
            
            current_row += 2
    
    # 调整列宽
    column_widths = {
        'A': 12, 'B': 15, 'C': 25, 'D': 20, 'E': 15,
        'F': 15, 'G': 15, 'H': 10, 'I': 10, 'J': 10, 'K': 30
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # 保存文件
    wb.save(output_path)
    
    return groups, df

